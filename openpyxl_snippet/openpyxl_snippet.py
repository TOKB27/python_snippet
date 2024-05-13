### Excel操作ライブラリのスニペット ###
import openpyxl

# Excelファイルの読み込み
wb = openpyxl.load_workbook('果物価格表.xlsx')

# シート名の取得
print(wb.sheetnames) # ['Sheet1']

# シートデータ取得
ws = wb["Sheet1"]

# データ取得1
value = ws["B3"].value

# データ取得2(以下はB3を表す)
value2 = ws.cell(row=3, column=2).value

# データがある場合のみ処理
if value2 is not None:
  print(f"value2:{value2}")

# セルの番地を指定してデータ取得(3〜5行目、2〜3列目の値を取り出す。)
for y in range(3, 6):
  r = []
  for x in range(2, 4):
    value = ws.cell(row=y, column=x).value
    r.append(value)
  print(r)

# セル名で範囲指定してデータ取得
for row in ws["B3:C5"]:
  print([c.value for c in row])

# 文字の色を取得
cell_text_color = ws["B3"].fill.fgColor.value
print(cell_text_color)

# 背景の色を取得
cell_bg_color = ws["B3"].fill.bgColor.value
print(cell_bg_color)


## グループ化状態の確認 ##
# チェックしたいセルの位置
cell_row, cell_column = 6, 2  # 例：5行目、2列目

# 行のグループ化状態を確認
row_grouped = ws.row_dimensions[cell_row].outline_level > 0

# 列のグループ化状態を確認
column_letter = openpyxl.utils.get_column_letter(cell_column)
column_grouped = ws.column_dimensions[column_letter].outline_level > 0

# 結果を表示
print(f"行 {cell_row} はグループ化されていますか？ {'はい' if row_grouped else 'いいえ'}")
print(f"列 {column_letter} はグループ化されていますか？ {'はい' if column_grouped else 'いいえ'}")